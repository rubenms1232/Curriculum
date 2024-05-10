import urllib.request
import json
import openpyxl
import random


def fetch_quote():
    # Fetch quotes from an API
    with urllib.request.urlopen("https://api.quotable.io/quotes/random?tags=health|gratitude|happiness|change|courage|ethics|failure|famous-quotes|future|freedom|inspirational") as response:
        quotes_data = response.read()
    quotes = json.loads(quotes_data.decode("utf-8"))

    # Select a random quote from the list
    selected_quote = random.choice(quotes)

    # Extract tags from the selected quote
    tags = ', '.join(selected_quote["tags"])

    return selected_quote["content"], selected_quote["author"], tags


# Create a new Excel workbook and add a worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Quotes"

# Add headers to the worksheet
worksheet.cell(row=1, column=1, value="Quote")
worksheet.cell(row=1, column=2, value="Author")
worksheet.cell(row=1, column=3, value="Tags")

# Generate and save 100 quotes
for i in range(2, 10002):  # Update the loop range to generate 100 quotes
    # Fetch a quote
    quote, author, tags = fetch_quote()

    # Write the quote, author, and tags to the worksheet
    worksheet.cell(row=i, column=1, value=quote)
    worksheet.cell(row=i, column=2, value=author)
    worksheet.cell(row=i, column=3, value=tags)

# Save the Excel file
workbook.save("quotes.xlsx")
